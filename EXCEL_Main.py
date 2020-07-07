
def Chicago():

    # This program identifies Chicago vendors who received TIF district subsidies in 2017 to 2018. It relies on a dataset
    # that lists TOTAL TIF payouts in the year of project completion. Therefore, we can determine which vendors have
    # recently completed development (or redevelopment) of their properties. We assume such properties have a high
    # likelihood of new or increased property taxes, and may be prime targets for the sales team.
    # We'll then see some broad conclusions from automated data analyses. NOTE: an analysis of this type will obviously
    # be more valuable with current data that Advantax may have access to, rather than data from 2 years ago.
    #
    # Next, I'll demonstrate mass data preparation in excel using a hypothetical scenario. Using the
    # format of our revised Chicago dataset we'll automate the creation of randomized, dummy data for 50 cities using TIF
    # districts. Each city will have its own spreadsheet, just like Chicago, saved for viewing.
    # Each new dataset will undergo the same process of data cleaning  and analyses. A list comprised of real company names
    # and 800+ automatically generated fictional companies will be used as random TIF money recipients.
    # Names of all non-Chicago TIF districts will be randomly generated with a huge dataset of street names.
    # Finally, a summary spreadsheet will be created showing the top recipients in the country, the top 5 cities
    # that spend the most on TIF districts, the top 5 highest funded individual districts, and the total of all TIF
    # district money spent by these 50 cities.
    # NOTE: this process is to demonstrate mass automation of data management: the randomized data is not accurate,
    # although I've fine-tuned some aspects of my randomization algorithm to produce realistic results.
    # Because the data is randomly generated, the results will be different each time I run the program.



    # Entire process will be a function. For Loop will be used in a master program that runs the function 50 times.
    # fun(input) will be city name, stored in master program list.

    ######################################## STEP 1:  Data Cleaning ###################################################

    # import needed modules and spreadsheets into Python VENV
    from sys import exit
    import random
    import pandas as pd
    import csv

    with open('C:\\Users\\Owner\\PycharmProjects\\ExcelMassManipulation\\venv\\TIF Vendor Payments (copy).csv', 'r') as sheet:
        X = csv.reader(sheet)
        TIF_payment_list = list(X)



    # record length of the spreadsheet for later use

    orig_len_payments = len(TIF_payment_list)
    if __name__ == "__main__":
        print(f'There are {orig_len_payments} entries in the Payments List for Chicago')
        print()
    # make copies of spreadsheets without headers for easier manipulation

    TIF_payment_2 = TIF_payment_list[:]
    del TIF_payment_2[0]

    # There are a wide variety of types of TIF payouts, and many of these go to vendors who do not own property. We need
    # to delete any data that isn't related to real estate development.

    desired_payment_types = ['Development', 'Acquisition', 'Rehabilitation Program']

    delcopy_TIF_payment2 = TIF_payment_2[:]
    if __name__ == "__main__":
        print('Now deleting all unwanted payment types in Chicago data set...')
    for row1 in delcopy_TIF_payment2:
        if row1[4] not in desired_payment_types:
            del_index1 = TIF_payment_2.index(row1)
            del TIF_payment_2[del_index1]
    if __name__ == "__main__":
        print('Deletion complete.')
        print()
        print('The revised data set:')
        print()
        print(TIF_payment_2)
        print()

    # calculation of total tif dollars spent on private real estate improvement

    payment_sum = 0.0
    for ind_pmt in TIF_payment_2:
        payment_sum += float(ind_pmt[5])

    payment_sum =int(payment_sum)
    if __name__ == "__main__":
        print(f'The sum of TIF payments Chicago made for private real estate improvement is ${payment_sum}')
        print()

    ### create list of vendors who received payment above $1,000,000. May be used to prioritize sales outreach.

    city_top_vendors = []

    for topvendor in TIF_payment_2:
        vendpayout = float(topvendor[5])
        if vendpayout >= 1000000.0:
            city_top_vendors.append(topvendor)

    ### identify top 5 TIF districts in the city. Could indicate rapidly growing property values and increased taxes within
    ### the district. Used to guide search for companies that may need professional tax help, whether they took subsidies
    ### or not.

    top_5 = []
    TIF_districts = []
    for district_add in TIF_payment_2:
        if district_add[1] not in TIF_districts:
            TIF_districts.append(district_add[1])

    district_and_sum = []
    for district in TIF_districts:
        district_index = TIF_districts.index(district)
        varrevised = []
        dis_total = 0.0
        for dist_search in TIF_payment_2:
            if dist_search[1] == district:
                dis_total += float(dist_search[5])
        varrevised.append(TIF_districts[district_index])
        varrevised.append(dis_total)
        district_and_sum.append(varrevised)

    # change the sums in the district and sum list from strings to integers. Necessary for calculations.
    for intconversion in district_and_sum:
        intconversion[1] = int(intconversion[1])
    if __name__ == "__main__":
        print('The full list of Chicago TIF districts with their subsidies is:')
    if __name__ == "__main__":
        for sd in district_and_sum:
            print(sd)
        print()

    # lists to be filled that will make top-5 calculations easier
    pmts_list = []
    top_5_pmts = []
    chicago_top5_districts = []

    # create list of all payments. It will make it much easier to find top 5 sums.
    for createpmtlist in district_and_sum:
        pmts_list.append(createpmtlist[1])

    # finding top 5 sums in the new payment list, add to top-5 pmt list
    pmts_list.sort(reverse=True)
    for it in range(5):
        top_5_pmts.append(pmts_list[it])

    if __name__ == "__main__":
        print(f'The top 5 payments made by Chicago are in the amount of : {top_5_pmts}')
        print()

    # match the top 5 amounts to their vendor

    for t5 in top_5_pmts:     # iterate through top 5 pmts
        for check in district_and_sum:    # look for each pmt within district and sum
            index_charlie = district_and_sum.index(check)
            if check[1] == t5:
                chicago_top5_districts.append(district_and_sum[index_charlie])
                break

    if __name__ == "__main__":
        print(f'The top 5 TIF districts in Chicago are {chicago_top5_districts}')

        print('chicago function imported correctly')
        print()

    # eliminate unneeded tif #s from master list
    for tifx in TIF_payment_2:
        del tifx[0]





    # find top 5 vendors
    chic_vendandpmts = []
    CHICAGO_TOP5_VENDS = []

    for iterv in TIF_payment_2:
        apd1 = iterv[2]
        apd2 = float(iterv[4])
        apd2 = int(apd2)
        chic_vendandpmts.append([apd1, apd2])

    if __name__ == "__main__":
        print('The list of Chicago Vendors and their payments is:')
        print()
        print(chic_vendandpmts)
        print()


    chicvend_nopmts = []
    rev_cvap = []


    # create a list of only the vendor names. Exclude repeats.
    for ita20 in chic_vendandpmts:
        if ita20[0] not in chicvend_nopmts:
            chicvend_nopmts.append(ita20[0])
    print()

    # for h in chicvend_nopmts:
    #     print(h)
    # print()

    ## make a list of the total spending per vendor
    for ita30 in chicvend_nopmts:
        running_tot_per_chivend = 0.0
        for nested_itera in chic_vendandpmts:
            if nested_itera[0] == ita30:
                running_tot_per_chivend += nested_itera[1]
        rev_cvap.append([ita30, running_tot_per_chivend])

    # for sl in rev_cvap:
    #     print(sl)
    print()

    # define a function that orders the revised vendors/payments from largest to smallest payment
    def sorty3(listy):
        # reverse = None (Sorts in Ascending order)
        # key is set to sort using second element of
        # sublist lambda has been used
        listy.sort(key=lambda x: x[1], reverse=True)
        return listy

    rev_cvap = sorty3(rev_cvap)

    # test the reordering
    # for hj in rev_cvap:
    #     print(hj)
    # print()

    # add the top 5 entries into the top5 vendor list
    for num50 in range(0, 5):
        CHICAGO_TOP5_VENDS.append(rev_cvap[num50])

    if __name__ == "__main__":
        print("The top 5 Vendors in Chicago are:")
        for printy40 in CHICAGO_TOP5_VENDS:
            print(printy40)
        print()

    # output of the function is the revised spreadsheet list
    return([TIF_payment_2, chicago_top5_districts, payment_sum, CHICAGO_TOP5_VENDS])




def topcities(city):
    import names
    import string
    import random as rd
    import openpyxl as xl
    from openpyxl.styles import Font, PatternFill, Color
    from openpyxl.styles.borders import Border, Side
    print()
    print(f'Next city is {city}')

    # use Chicago dataset as template to replace with new random data
    chicreturn = Chicago()
    new_data_set = chicreturn[0]

    print()

    # replace tif sums with new random number
    for rep1 in new_data_set:
        indexDS = new_data_set.index(rep1)
        rx = rd.randint(0,25)
        if rx in range(0,22):
            r = rd.randint(10000, 150000)
        elif rx in range(22, 25):
            r = rd.randint(50000, 2200000)
        else:
            r = rd.randint(0, 21000000)
        new_data_set[indexDS][4] = r

    # replace type of tif payment with random type
    for rep2 in new_data_set:
        r2 = rd.randint(0,10)
        if r2 in range(9,11):
            tif_type = ['Acquisition']
        else:
            tif_type = ['Development','Rehabilitation Program']
        tif_selection = rd.choice(tif_type)
        indexDS2 = new_data_set.index(rep2)
        new_data_set[indexDS2][3] = tif_selection


    # use a series of random generations to fill in the vender column

    # def vendorgen():
    #
    #     company_type = ['Co', 'Corps', 'LLC', 'Partners', 'Pension Fund', 'Hedge Fund', 'Company', 'International', 'Global'
    #                     'Incorporated', 'Systems', 'Industries', 'Services', 'Group', 'Technologies', 'Investments',
    #                     'Insurance Co.', 'Foundation', 'Trust', 'Development', 'Developers', 'Rehab Specialists',
    #                     'Institute']
    #
    #     alphabet_list = list(string.ascii_uppercase)

        name_list = []

        # This is the name generator that was used to create the 800+ list of fictional companies.
        # It can be used in the future for a new list, but it runs slowly. Program runs much faster with a list
        # defined before runtime and without use of random module.


        # for iter in range(0, 100):
        #     rxx = rd.randint(0,100)
        #     if rxx in range(0, 20):
        #         new_name = names.get_last_name() + " & " + names.get_last_name()
        #         name_list.append(new_name)
        #     elif rxx in range(20, 40):
        #         letter = rd.choice(alphabet_list)
        #         new_name = names.get_first_name('male') + " " + letter + " " + names.get_last_name()
        #         name_list.append(new_name)
        #     elif rxx in range(40, 60):
        #         letter = rd.choice(alphabet_list)
        #         letter2 = rd.choice(alphabet_list)
        #         new_name = f'{letter}.{letter2}. {names.get_last_name()}'
        #         name_list.append(new_name)
        #     else:
        #         new_name = names.get_full_name('male')
        #         name_list.append(new_name)
        #
        #
        # for rnum1 in range(0,300):
        #     vendor_name = (f'{rd.choice(name_list)} {rd.choice(company_type)}')


        # return vendor_name

    existing_RE_companies = ['AvalonBay Communities', 'Greystar Real Estate Partners', 'Wood Partners',
                             'Mill Creek Residential', 'Continental Properties Company, Inc.', 'Hines',
                             'Trammell Crow Company', 'The JBG Companies', 'Lowe Enterprises', 'Hillwood',
                             'Simon Property Group', 'General Growth Properties', 'SITE Centers', 'Kimco Realty Corp',
                             'Brixmor Property Group', 'ProLogis', 'Panattoni Development Co.',
                             ' McDonald Development Co.', 'USAA Real Estate Co.', 'LaSalle Investment Management',
                             'Lincoln Property Company', 'Alliance Residential', 'JPI', 'The Michaels Organization',
                             'American Tower Corporation', 'Annaly Capital Management', 'Weyerhaeuser', 'CBRE Group'
                                                                                                        'Ventas',
                             'Host Hotels & Resorts', 'Equinix', 'Welltower', 'Digital Realty Trust',
                             'Two Harbors Investment', 'Chimera Investment', 'Vornado Realty', 'Realty Income',
                             'Essex Property Trust', 'Alexandria Real Estate Equities', 'HCP', 'Prudential',
                             'United Technologies', 'Tesla Motors', 'Apple', 'Microsoft, Inc.', 'Morgan Stanley',
                             'J.P. Morgan Chase', 'Nationwide', 'Tyson Foods', 'World Fuel Services', 'Progressive',
                             'Geico Insurance', 'Capital One Financial', 'Exelon', 'Publix', 'Meijer - Dev. Division',
                             'Northrop Gruman', 'Phillip Morris Internationl', 'TY Lin International',
                             'Aecom Technical Service', 'USDA', 'Raytheon', 'NewCorps', 'PBF Construction',
                             'Jewel Food Stores', 'CVS', 'Target Corps.', 'Khols', 'Bechtel Group Inc',
                             'Fluor Corporation', 'Jacobs Engineering Group INC.', 'Kiewit Corporation',
                             'Turner Construction', 'Skanska Construction', 'Poole Construction Ltd ', 'Walmart',
                             'Hilton', 'Wegmans Food Markets', 'Edward Jones', 'American Express',
                             'Pinaccle Financial Partners', 'Navy Federal Credit Union', 'Carmax', 'Nvidia',
                             'David Weekley Homes']


    vendorlist = ['Timothy C Tandy Insurance Co.', 'George Sullivan Hedge Fund', 'Wayne Antonucci Trust', 'Z.Q. Hoffmann Corps', 'Johnson & Hoover Rehab Specialists', 'X.W. Davis Systems', 'Furlow & Eyre LLC', 'Benjamin A Smith Co', 'S.U. Mayes Development', 'Howard S Walk Technologies', 'Minarik & Lopez Foundation', 'Rafael P Sears Development', 'Thomas Anderson LLC', 'Arthur Wells Hedge Fund', 'I.F. Warrick Services', 'Joseph Ramos Company', 'George Cole Rehab Specialists', 'Todd Stammer Trust', 'L.U. Mcmullen Corps', 'Young & Rhodes Hedge Fund', 'Jacques & Hoffman Group', 'Darin Mcqueen Company', 'James I Lee Corps', 'Kevin D Fitzpatrick Corps', 'L.O. Fuhrman Development', 'Frank Syvertsen Rehab Specialists', 'Canty & Pryor Corps', 'Paul V Kinkead Services', 'Harold V Thompson Pension Fund', 'Faunce & Zimmer Partners', 'Lowe & Hunter Industries', 'Charles F Lloyd Development', 'Brian Whiting Systems', 'Reginald West Services', 'Patrick K Draggoo Development', 'Reed & Cirillo Development', 'Nickel & Herandez Trust', 'Roy Handley Foundation', 'Perez & Schmitz Group', 'Remley & Snyder Partners', 'Y.O. Harris GlobalIncorporated', 'Troy Page Company', 'Fisher & Myers Foundation', 'Antonio Martin Corps', 'Donald Murray Trust', 'N.G. Collins Insurance Co.', 'Salvador Washington Group', 'Brown & Hurtado LLC', 'I.E. Reed Foundation', 'R.F. Ward Technologies', 'Simon & Simmons Partners', 'Heuser & Bokor Trust', 'Lee Desrochers Foundation', 'Tom Morris Industries', 'Joseph H Zappulla Investments', 'Brooks & Burton Technologies', 'Buchholz & Denoon Group', 'J.X. Burns Developers', 'Michael Lang Development', 'Walter D Booth International', 'Bryan Stimmel LLC', 'A.K. Mccann Insurance Co.', 'Saunders & Mcsherry Technologies', 'Ashley H Leaton Hedge Fund', 'Thomas D Button Corps', 'James R Lundgren Institute', 'Harold Mullins Hedge Fund', 'Scott Vasquez Investments', 'Lewis Brand Investments', 'B.K. Licea Systems', 'John Landfried Services', 'Robert Schroeder International', 'D.U. Loughry Services', 'James U Youngblood Development', 'Calvin M East Group', 'Ernest Gabrielson Company', 'Norman Baldwin Industries', 'Johnny M Dougherty Insurance Co.', 'Jeff Johns Insurance Co.', 'O.I. Charland LLC', 'Michael Quintero Corps', 'Alan N Ator Industries', 'L.H. Edwards Corps', 'Kevin N Henshaw Industries', 'James & Bowling Industries', 'Luis O Brooks Investments', 'T.I. Ochoa LLC', 'U.J. Potter Corps', 'M.S. Muncy Hedge Fund', 'Bobby B Murdock LLC', 'P.T. Rogers Company', 'Dale Mangrum Corps', 'Ramirez & Rodriquez Pension Fund', 'J.I. Pollard Rehab Specialists', 'Henry A Cruz Industries', 'O.U. Chew Development', 'William Mcdaries Development', 'D.F. Garcia Investments', 'David Steese Pension Fund', 'Hillard & Tijerina International', 'Ken A Downey Development', 'Kaufman & Hart Corps', 'Abel Castaneda Group', 'David X Quarles GlobalIncorporated', 'Jason Rivera Pension Fund', 'Roberto & Ramaker Hedge Fund', 'David D Boccella Hedge Fund', 'W.O. Glenn Technologies', 'A.W. Johnson International', 'Phillips & Dixon Institute', 'Alfred Monson Hedge Fund', 'Shawn Dixon International', 'Lloyd Neely Trust', 'John Jackson Group', 'Melvin N Speece Pension Fund', 'O.J. Day Rehab Specialists', 'James B Ramsey Services', 'Raymond Sell Development', 'Caldwell & Camacho Company', 'Jon Barker Partners', 'Phipps & Sanchez Institute', 'Robles & Dorey Group', 'Ramos & Jenkins International', 'A.O. Sadowski Development', 'Jerry Cotman International', 'J.D. Bakken Company', 'Harry Stephens Insurance Co.', 'John Myers Investments', 'W.Q. Jones LLC', 'I.W. Waters Developers', 'Griffin & Sexton Pension Fund', 'Ramirez & Simpson International', 'M.V. Faughnan Technologies', 'Brad T Friedman Foundation', 'Edward Clark GlobalIncorporated', 'Clyde Freedman Rehab Specialists', 'Lionel A Toney Developers', 'Sean Forsyth Rehab Specialists', 'Luciano Sala Rehab Specialists', 'William B Wheeler Foundation', 'Wiley & Birkline Rehab Specialists', 'Longstreet & Hitchcock Pension Fund', 'John Sodomka Development', 'D.G. Campbell Technologies', 'Maurice S Buck Industries', 'Rogelio N Neely International', 'Randy Oakley Group', 'Richard C Wilson Services', 'William Anna Development', 'U.Y. Collier Systems', 'Kelvin Y Fujii Institute', 'X.T. Baker Partners', 'Michael X Benware Co', 'G.Z. Knoll Technologies', 'M.L. Coulombe Industries', 'Blake Hill Rehab Specialists', 'Palms & Thompson Development', 'Trey Durham Pension Fund', 'Jason Q Quinter GlobalIncorporated', 'D.S. Ainsworth Institute', 'Christensen & Villaman Corps', 'Alexander Sutton Hedge Fund', 'Hoffman & Debellis Rehab Specialists', 'Gill & Thompson GlobalIncorporated', 'R.T. Burns Partners', 'Hvizdos & Shuman Developers', 'Thomas K Wilson Trust', 'Lewis Cumbee Investments', 'Nick Almaguer Developers', 'T.P. Williams Institute', 'Z.O. Telfer Co', 'Ted Y Bushovisky Technologies', 'Mckinley & Smiley Pension Fund', 'Gay & Addleman Investments', 'John Brown Co', 'X.V. Sarracino Foundation', 'E.E. Vaughn Investments', 'William Littlewood Institute', 'Robert Sapp Institute', 'Samuel Richardson Trust', 'Mark Crossmon Company', 'Robert Cota Development', 'X.Q. Pollack Developers', 'James Coleman Services', 'James Kearney Technologies', 'Preston & Wilson Insurance Co.', 'Kenneth U Schenk Technologies', 'Michael Smith Services', 'C.J. Jimenez Technologies', 'Kenneth Mccoy LLC', 'Fredrick B Byrd Company', 'Willie M Osbourne LLC', 'Timothy I Henry Technologies', 'John Spiller Trust', 'R.D. Nunemaker GlobalIncorporated', 'Thomas Collins Partners', 'Daniel Bilbro Rehab Specialists', 'Leighton & Taylor Development', 'Bryan E Buisson Rehab Specialists', 'Fred Hayes Technologies', 'Barry Oland Industries', 'Donald Marcum Co', 'Mullikin & Chavis GlobalIncorporated', 'Mark Taylor International', 'Vernon Bunn Industries', 'B.M. Croteau Industries', 'Parker & Crawford Group', 'Matthew I Garza Rehab Specialists', 'Cardin & Terry LLC', 'O.L. Jackson Services', 'Timothy Harvey Corps', 'Cruz & Mcclellan Pension Fund', 'Y.Z. Cowan Developers', 'Louis R Carlson Systems', 'Arthur Stafford Hedge Fund', 'David Dardis Industries', 'Frank Corkins Group', 'Arnold G Mclean LLC', 'Francisco Vantassel Insurance Co.', 'James K Strawser Investments', 'H.N. Dykstra Institute', 'Raymond W Sallade Institute', 'Robby O Brickell Systems', 'Elliot Greer Corps', 'R.C. Randall Pension Fund', 'C.X. Surratt Rehab Specialists', 'Christopher Yin Corps', 'John Wilson LLC', 'Michael Salmon Co', 'U.J. Newson Corps', 'Donald Champion Trust', 'John Mosby Institute', 'Mercer & Miles LLC', 'Lunde & Hopkins Corps', 'W.O. Ballard Insurance Co.', 'Corliss & Fraser Institute', 'Ronald Nagle Pension Fund', 'Robert V Locke Insurance Co.', 'Jeffrey V Smith International', 'Gordon & Palmore Group', 'P.A. Switzer Insurance Co.', 'Thomas Gonzalez Foundation', 'I.M. Vahey Institute', 'Bennie Nelson Co', 'I.G. Adams Hedge Fund', 'Terry Askew Rehab Specialists', 'Benjamin V Lipe Corps', 'F.Q. Fell LLC', 'Steven I Rhodd Partners', 'D.Z. Walker Co', 'Andrew O Abel Investments', 'Mark Shapiro Services', 'Kenneth Sallade Group', 'R.A. Anderson Development', 'Ernest Moyer GlobalIncorporated', 'Richard Lee Co', 'Olsen & Nelson Trust', 'Paul Solano Technologies', 'David X Parker Pension Fund', 'Lund & Story Industries', 'Aggas & Oneill Rehab Specialists', 'G.H. Wiggains Investments', 'H.R. Pleva GlobalIncorporated', 'Gerald Gaynor Foundation', 'O.R. Schaich Systems', 'Keith Laflam Pension Fund', 'Bruce Johnson LLC', 'I.X. Green Insurance Co.', 'Jose L Werkhoven Industries', 'Charles Haines Industries', 'Hayden & Barnes Insurance Co.', 'Bleiweiss & Holmes Trust', 'Raymond Fitzgibbon Services', 'John Dyer Development', 'W.G. Clayton Pension Fund', 'Louis Effron Pension Fund', 'David Cuccia International', 'B.G. Williams GlobalIncorporated', 'Federico Eaton Developers', 'William B Wright Developers', 'Clarence Stewart International', 'Craig C Cash Trust', 'Charles Acosta Institute', 'C.I. Mcelrath Industries', 'Joseph Diehl Group', 'Robert Kimball Rehab Specialists', 'Daryl Fletcher GlobalIncorporated', 'Penn & Creel Group', 'N.X. Kelly Technologies', 'James Jones Insurance Co.', 'Marvin P Schilling Industries', 'Christopher Z Hoelscher Company', 'Jones & Hill Co', 'Anthony B Starbuck Company', 'I.B. Brown Institute', 'Randall Loughran Trust', 'John E Benavides Group', 'K.Q. Mathis LLC', 'Andy Smith Technologies', 'Kirk G Callahan International', 'Ryan Garcia LLC', 'Orlando Miller Hedge Fund', 'Steven Mack Trust', 'Jason Blasen LLC', 'Richard F Williams Company', 'Gary Raffa Services', 'N.L. Reeves Partners', 'J.L. Baltes Systems', 'Jesse Q Pichette Institute', 'R.J. Dean Company', 'Lecky & Mcdonnell Technologies', 'Bump & Wynn Investments', 'Ronald Gottschalk Industries', 'M.O. Myers Services', 'C.K. Abbott LLC', 'Bramuchi & Moore Co', 'A.Y. Blackburn Developers', 'Bernard Williams Technologies', 'Joseph Barber Rehab Specialists', 'Banks & Lietzke Services', 'Y.O. Graham Technologies', 'Eldon W Healey Trust', 'Gonzalez & Schneider International', 'R.U. Graffagnino Partners', 'John V Atchley Foundation', 'Anthony U Callahan Foundation', 'Stokes & Stitt LLC', 'Michael Vierra Co', 'Parker & Bunch Hedge Fund', 'Mowery & Belser Group', 'Theodore G Poindexter Services', 'Marty & Hanson Investments', 'Farrell & Harrison Institute', 'Humberto Renfro Company', 'Walker & Thomas Insurance Co.', 'Robert Gutierrez Corps', 'Ken Sandell Corps', 'David Field Partners', 'Harold Sumner GlobalIncorporated', 'Michael Rivera Institute', 'Mccall & Bagby GlobalIncorporated', 'Q.Y. Mona GlobalIncorporated', 'Dallas Carrillo Co', 'Matthew Jackson Trust', 'Roger P Foster Insurance Co.', 'Baber & Baker Industries', 'Terry Crawford Developers', 'Gelsinger & Gallant Trust', 'David Varella Investments', 'Eric S Meinzer Industries', 'Alexander V Seavey Trust', 'Hosea Hughes Systems', 'Michael Jones Trust', 'Thomas & Niebuhr Company', 'James & Johnson Insurance Co.', 'S.U. Jessop Institute', 'John Ranieri International', 'Woodrow D Ruffo Rehab Specialists', 'E.F. Cerny Rehab Specialists', 'A.G. Lockwood Rehab Specialists', 'Richard Headley Company', 'Murphy & Carrier Pension Fund', 'W.Y. Hirsch Development', 'David Boren Services', 'Reyes Ouellette Foundation', 'Terrance Samples Systems', 'Robert U Girard Insurance Co.', 'Joseph Rucker Rehab Specialists', 'Q.O. Willie Technologies', 'James Saunders Trust', 'Clarence Pack Industries', 'Miller & Mckee Services', 'Clemente Walls Co', 'Michael Werner Rehab Specialists', 'Daniel & Pena Partners', 'Jacob Blake Investments', 'Wayne Valliere Services', 'U.V. Burger Services', 'W.T. Withem Services', 'Fred Abad Development', 'O.E. Scott Developers', 'Byron Harpin International', 'Adrian Q Carlson Company', 'William Aronson Co', 'Sutton & Christensen Trust', 'Leeds & Schmidt Insurance Co.', 'Charles Ford Institute', 'Noel & Bartberger Pension Fund', 'Evan Glaab Systems', 'M.Q. Weathers Trust', 'Henry Roberts Corps', 'Aaron Buckner Partners', 'William Hernandez Pension Fund', 'Tracy Yamamoto Technologies', 'Joseph Merchant Systems', 'F.T. Johnson Services', 'Pollick & Borrego Foundation', 'Thompson & Cooper Company', 'Harris & Reando Company', 'Ernest D Shively Services', 'I.F. Warrick Services', 'Joseph Ramos Company', 'George Cole Rehab Specialists', 'Todd Stammer Trust', 'L.U. Mcmullen Corps', 'Young & Rhodes Hedge Fund', 'Jacques & Hoffman Group', 'Darin Mcqueen Company', 'James I Lee Corps', 'Kevin D Fitzpatrick Corps', 'L.O. Fuhrman Development', 'Frank Syvertsen Rehab Specialists', 'Canty & Pryor Corps', 'Paul V Kinkead Services', 'Harold V Thompson Pension Fund', 'Faunce & Zimmer Partners', 'Lowe & Hunter Industries', 'Charles F Lloyd Development', 'Brian Whiting Systems', 'Reginald West Services', 'Patrick K Draggoo Development', 'Reed & Cirillo Development', 'Nickel & Herandez Trust', 'Roy Handley Foundation', 'Perez & Schmitz Group', 'Remley & Snyder Partners', 'Y.O. Harris GlobalIncorporated', 'Troy Page Company', 'Fisher & Myers Foundation', 'Antonio Martin Corps', 'Donald Murray Trust', 'N.G. Collins Insurance Co.', 'Salvador Washington Group', 'Brown & Hurtado LLC', 'I.E. Reed Foundation', 'R.F. Ward Technologies', 'Simon & Simmons Partners', 'Heuser & Bokor Trust', 'Lee Desrochers Foundation', 'Tom Morris Industries', 'Joseph H Zappulla Investments', 'Brooks & Burton Technologies', 'Buchholz & Denoon Group', 'J.X. Burns Developers', 'Michael Lang Development', 'Walter D Booth International', 'Bryan Stimmel LLC', 'A.K. Mccann Insurance Co.', 'Saunders & Mcsherry Technologies', 'Ashley H Leaton Hedge Fund', 'Thomas D Button Corps', 'James R Lundgren Institute', 'Harold Mullins Hedge Fund', 'Scott Vasquez Investments', 'Lewis Brand Investments', 'B.K. Licea Systems', 'John Landfried Services', 'Robert Schroeder International', 'D.U. Loughry Services', 'James U Youngblood Development', 'Calvin M East Group', 'Ernest Gabrielson Company', 'Norman Baldwin Industries', 'Johnny M Dougherty Insurance Co.', 'Jeff Johns Insurance Co.', 'O.I. Charland LLC', 'Michael Quintero Corps', 'Alan N Ator Industries', 'L.H. Edwards Corps', 'Kevin N Henshaw Industries', 'James & Bowling Industries', 'Luis O Brooks Investments', 'T.I. Ochoa LLC', 'U.J. Potter Corps', 'M.S. Muncy Hedge Fund', 'Bobby B Murdock LLC', 'P.T. Rogers Company', 'Dale Mangrum Corps', 'Ramirez & Rodriquez Pension Fund', 'J.I. Pollard Rehab Specialists', 'Henry A Cruz Industries', 'O.U. Chew Development', 'William Mcdaries Development', 'D.F. Garcia Investments', 'David Steese Pension Fund', 'Hillard & Tijerina International', 'Ken A Downey Development', 'Kaufman & Hart Corps', 'Abel Castaneda Group', 'David X Quarles GlobalIncorporated', 'Jason Rivera Pension Fund', 'Roberto & Ramaker Hedge Fund', 'David D Boccella Hedge Fund', 'W.O. Glenn Technologies', 'A.W. Johnson International', 'Phillips & Dixon Institute', 'Alfred Monson Hedge Fund', 'Shawn Dixon International', 'Lloyd Neely Trust', 'John Jackson Group', 'Melvin N Speece Pension Fund', 'O.J. Day Rehab Specialists', 'James B Ramsey Services', 'Raymond Sell Development', 'Caldwell & Camacho Company', 'Jon Barker Partners', 'Phipps & Sanchez Institute', 'Robles & Dorey Group', 'Ramos & Jenkins International', 'A.O. Sadowski Development', 'Jerry Cotman International', 'J.D. Bakken Company', 'Harry Stephens Insurance Co.', 'John Myers Investments', 'W.Q. Jones LLC', 'I.W. Waters Developers', 'Griffin & Sexton Pension Fund', 'Ramirez & Simpson International', 'M.V. Faughnan Technologies', 'Brad T Friedman Foundation', 'Edward Clark GlobalIncorporated', 'Clyde Freedman Rehab Specialists', 'Lionel A Toney Developers', 'Sean Forsyth Rehab Specialists', 'Luciano Sala Rehab Specialists', 'William B Wheeler Foundation', 'Wiley & Birkline Rehab Specialists', 'Longstreet & Hitchcock Pension Fund', 'John Sodomka Development', 'D.G. Campbell Technologies', 'Maurice S Buck Industries', 'Rogelio N Neely International', 'Randy Oakley Group', 'Richard C Wilson Services', 'William Anna Development', 'U.Y. Collier Systems', 'Kelvin Y Fujii Institute', 'X.T. Baker Partners', 'Michael X Benware Co', 'G.Z. Knoll Technologies', 'M.L. Coulombe Industries', 'Blake Hill Rehab Specialists', 'Palms & Thompson Development', 'Trey Durham Pension Fund', 'Jason Q Quinter GlobalIncorporated', 'D.S. Ainsworth Institute', 'Christensen & Villaman Corps', 'Alexander Sutton Hedge Fund', 'Hoffman & Debellis Rehab Specialists', 'Gill & Thompson GlobalIncorporated', 'R.T. Burns Partners', 'Hvizdos & Shuman Developers', 'Thomas K Wilson Trust', 'Lewis Cumbee Investments', 'Nick Almaguer Developers', 'T.P. Williams Institute', 'Z.O. Telfer Co', 'Ted Y Bushovisky Technologies', 'Mckinley & Smiley Pension Fund', 'Gay & Addleman Investments', 'John Brown Co', 'X.V. Sarracino Foundation', 'E.E. Vaughn Investments', 'William Littlewood Institute', 'Robert Sapp Institute', 'Samuel Richardson Trust', 'Mark Crossmon Company', 'Robert Cota Development', 'X.Q. Pollack Developers', 'James Coleman Services', 'James Kearney Technologies', 'Preston & Wilson Insurance Co.', 'Kenneth U Schenk Technologies', 'Michael Smith Services', 'C.J. Jimenez Technologies', 'Kenneth Mccoy LLC', 'Fredrick B Byrd Company', 'Willie M Osbourne LLC', 'Timothy I Henry Technologies', 'John Spiller Trust', 'R.D. Nunemaker GlobalIncorporated', 'Thomas Collins Partners', 'Daniel Bilbro Rehab Specialists', 'Leighton & Taylor Development', 'Bryan E Buisson Rehab Specialists', 'Fred Hayes Technologies', 'Barry Oland Industries', 'Donald Marcum Co', 'Mullikin & Chavis GlobalIncorporated', 'Mark Taylor International', 'Vernon Bunn Industries', 'B.M. Croteau Industries', 'Parker & Crawford Group', 'Matthew I Garza Rehab Specialists', 'Cardin & Terry LLC', 'O.L. Jackson Services', 'Timothy Harvey Corps', 'Cruz & Mcclellan Pension Fund', 'Y.Z. Cowan Developers', 'Louis R Carlson Systems', 'Arthur Stafford Hedge Fund', 'David Dardis Industries', 'Frank Corkins Group', 'Arnold G Mclean LLC', 'Francisco Vantassel Insurance Co.', 'James K Strawser Investments', 'H.N. Dykstra Institute', 'Raymond W Sallade Institute', 'Robby O Brickell Systems', 'Elliot Greer Corps', 'R.C. Randall Pension Fund', 'C.X. Surratt Rehab Specialists', 'Christopher Yin Corps', 'John Wilson LLC', 'Michael Salmon Co', 'U.J. Newson Corps', 'Donald Champion Trust', 'John Mosby Institute', 'Mercer & Miles LLC', 'Lunde & Hopkins Corps', 'W.O. Ballard Insurance Co.', 'Corliss & Fraser Institute', 'Ronald Nagle Pension Fund', 'Robert V Locke Insurance Co.', 'Jeffrey V Smith International', 'Gordon & Palmore Group', 'P.A. Switzer Insurance Co.', 'Thomas Gonzalez Foundation', 'I.M. Vahey Institute', 'Bennie Nelson Co', 'I.G. Adams Hedge Fund', 'Terry Askew Rehab Specialists', 'Benjamin V Lipe Corps', 'F.Q. Fell LLC', 'Steven I Rhodd Partners', 'D.Z. Walker Co', 'Andrew O Abel Investments', 'Mark Shapiro Services', 'Kenneth Sallade Group', 'R.A. Anderson Development', 'Ernest Moyer GlobalIncorporated', 'Richard Lee Co', 'Olsen & Nelson Trust', 'Paul Solano Technologies', 'David X Parker Pension Fund', 'Lund & Story Industries', 'Aggas & Oneill Rehab Specialists', 'G.H. Wiggains Investments', 'H.R. Pleva GlobalIncorporated', 'Gerald Gaynor Foundation', 'O.R. Schaich Systems', 'Keith Laflam Pension Fund', 'Bruce Johnson LLC', 'I.X. Green Insurance Co.', 'Jose L Werkhoven Industries', 'Charles Haines Industries', 'Hayden & Barnes Insurance Co.', 'Bleiweiss & Holmes Trust', 'Raymond Fitzgibbon Services', 'John Dyer Development', 'W.G. Clayton Pension Fund', 'Louis Effron Pension Fund', 'David Cuccia International', 'B.G. Williams GlobalIncorporated', 'Federico Eaton Developers', 'William B Wright Developers', 'Clarence Stewart International', 'Craig C Cash Trust', 'Charles Acosta Institute', 'C.I. Mcelrath Industries', 'Joseph Diehl Group', 'Robert Kimball Rehab Specialists', 'Daryl Fletcher GlobalIncorporated', 'Penn & Creel Group', 'N.X. Kelly Technologies', 'James Jones Insurance Co.', 'Marvin P Schilling Industries', 'Christopher Z Hoelscher Company', 'Jones & Hill Co', 'Anthony B Starbuck Company', 'I.B. Brown Institute', 'Randall Loughran Trust', 'John E Benavides Group', 'K.Q. Mathis LLC', 'Andy Smith Technologies', 'Kirk G Callahan International', 'Ryan Garcia LLC', 'Orlando Miller Hedge Fund', 'Steven Mack Trust', 'Jason Blasen LLC', 'Richard F Williams Company', 'Gary Raffa Services', 'N.L. Reeves Partners', 'J.L. Baltes Systems', 'Jesse Q Pichette Institute', 'R.J. Dean Company', 'Lecky & Mcdonnell Technologies', 'Bump & Wynn Investments', 'Ronald Gottschalk Industries', 'M.O. Myers Services', 'C.K. Abbott LLC', 'Bramuchi & Moore Co', 'A.Y. Blackburn Developers', 'Bernard Williams Technologies', 'Joseph Barber Rehab Specialists', 'Banks & Lietzke Services', 'Y.O. Graham Technologies', 'Eldon W Healey Trust', 'Gonzalez & Schneider International', 'R.U. Graffagnino Partners', 'John V Atchley Foundation', 'Anthony U Callahan Foundation', 'Stokes & Stitt LLC', 'Michael Vierra Co', 'Parker & Bunch Hedge Fund', 'Mowery & Belser Group', 'Theodore G Poindexter Services', 'Marty & Hanson Investments', 'Farrell & Harrison Institute', 'Humberto Renfro Company', 'Walker & Thomas Insurance Co.', 'Robert Gutierrez Corps', 'Ken Sandell Corps', 'David Field Partners', 'Harold Sumner GlobalIncorporated', 'Michael Rivera Institute', 'Mccall & Bagby GlobalIncorporated', 'Q.Y. Mona GlobalIncorporated', 'Dallas Carrillo Co', 'Matthew Jackson Trust', 'Roger P Foster Insurance Co.', 'Baber & Baker Industries', 'Terry Crawford Developers', 'Gelsinger & Gallant Trust', 'David Varella Investments', 'Eric S Meinzer Industries', 'Alexander V Seavey Trust', 'Hosea Hughes Systems', 'Michael Jones Trust', 'Thomas & Niebuhr Company', 'James & Johnson Insurance Co.', 'S.U. Jessop Institute', 'John Ranieri International', 'Woodrow D Ruffo Rehab Specialists', 'E.F. Cerny Rehab Specialists', 'A.G. Lockwood Rehab Specialists', 'Richard Headley Company', 'Murphy & Carrier Pension Fund', 'W.Y. Hirsch Development', 'David Boren Services', 'Reyes Ouellette Foundation', 'Terrance Samples Systems', 'Robert U Girard Insurance Co.', 'Joseph Rucker Rehab Specialists', 'Q.O. Willie Technologies', 'James Saunders Trust', 'Clarence Pack Industries', 'Miller & Mckee Services', 'Clemente Walls Co', 'Michael Werner Rehab Specialists', 'Daniel & Pena Partners', 'Jacob Blake Investments', 'Wayne Valliere Services', 'U.V. Burger Services', 'W.T. Withem Services', 'Fred Abad Development', 'O.E. Scott Developers', 'Byron Harpin International', 'Adrian Q Carlson Company', 'William Aronson Co', 'Sutton & Christensen Trust', 'Leeds & Schmidt Insurance Co.', 'Charles Ford Institute', 'Noel & Bartberger Pension Fund', 'Evan Glaab Systems', 'M.Q. Weathers Trust', 'Henry Roberts Corps', 'Aaron Buckner Partners', 'William Hernandez Pension Fund', 'Tracy Yamamoto Technologies', 'Joseph Merchant Systems', 'F.T. Johnson Services', 'Pollick & Borrego Foundation', 'Thompson & Cooper Company', 'Harris & Reando Company', 'Ernest D Shively Services']
    final_vendor_list = vendorlist + existing_RE_companies

    for vender_iter in new_data_set:
        index_iter = new_data_set.index(vender_iter)
        new_data_set[index_iter][2] = rd.choice(final_vendor_list)

    street_selector = ['McKibbin/Lonsdale', 'George/Green Ridge', 'Essen/Hearle', 'Evana/Alyce', 'Lynnwood/Leonardo', 'E Lawn/Barrie', 'Attridge/Filmore', 'Milliken/Oxbow', 'Hortense/Windemere', 'Southwood/Carrie', 'McGee/Burro', 'Old Britton/Bataan', 'Packard/Windy', 'Shirra/Gosling Hill', 'Yarnell/Stony Hill', 'O Neil/Fireside', 'Chadwick Oaks/Hett', 'W Cambridge/Alps', 'Aske/Nurge', 'Brinley/New Walnut', 'Hudson Park/Schwinn', 'Brownell/Mc Coy', 'Grenfell/E Broadway', 'S Sunnyside/Tropical', 'Reni/Lange', 'Palisades Center/Marcus Garvey', 'Bouck/South Airmont', 'Chemka Pool/West Lawn', 'Sperry/Aske', 'East Crescent/Verdi', 'Brasch/Marcus Garvey', 'Merrall/Mc Coy', 'Jeanna/Bailiwick', 'Burro/Airport', 'Brambach/Radstock', 'Cliffdale/Fuhrman', 'E Inman/Heatherwood', 'Portside/Lysbeth', 'Marsh/Shaw', 'Wanamaker/Currans', 'Brownell/Roundtop', 'Fairmead/Irving', 'Embry Farm/Hardwell', 'Blue Spruce/Knowles', 'Fortesque/Appleby', 'Landry/Red Cross', 'S Beechcroft/Helmetta Jamesburg', 'Netherwood/Filmore', 'East Elbrook/Tailor', 'W Oxford/Southwood', 'McKibbin/Highbrook', 'Stokes Farm/Clearmeadow', 'Hadfield/Club', 'Twin Oaks/Stivers', 'Bear/S Beechcroft', 'Helmetta Jamesburg/Von Huenfeld', 'W Wheatley/Herb Hill', 'Twin Oak/Alps', 'Booker/Embry Farm', 'Sleepy Hollow/Saint George', 'Hoyts/Crosspointe', 'Jeanna/Wanamaker', 'Alps/Paddington', 'Sleepy Hollow/Carol', 'N Niagra/Elmer', 'Lohman/Green Ridge', 'Yarnell/Seaton', 'Edi/E Elizabeth', 'Blue Spruce/Appleby', 'Normalee/Alize', 'Mullen/N Bradford', 'Abington/Dixon', 'Old Field Point/Barrie', 'Interlaken/McKibbin', 'Cherokee/Marsh', 'Ruby/W Cambridge', 'Chadwick Oaks/Esterbrook', 'Mc Kee/Hawkstone', 'Hyannis/Milliken', 'Allen/Cheyenne', 'Ellin/Kahl', 'Alger/Bush', 'N Haledon/John F Kennedy', 'Milburn/Cordwood', 'Tailor/Algiers', 'W Central/Scudder', 'Alps/Bellrose', 'Berrybush/Curzon', 'Cherokee/E Broad', 'N Washington/Peerless', 'Lamar/Round', 'Arkay/Fuhrman', 'O Neil/Club', 'E Cambridge/Portside', 'Borinsky/Saint George', 'Foxhunt/N Bayview', 'Brownell/Hawkstone', 'Premier/Swannekin', 'Needes/Brae Burn', 'Fuhrman/Cos Cob', 'Merrall/Jeanna', 'Hyannis/Alize', 'Webster Valley/Alyce', 'Yarnell/Abraham', 'Shirra/Jagoe', 'Brinley/N Regent', 'Dallas/Deseret', 'Soloff/Currans', 'Rever/Hancock', 'Algiers/Waltoffer', 'E Birch/Brookshire', 'E Peddie/Elias', 'Attridge/S Wood', 'Allegany/Michael F', 'Vernon/S Corona', 'Newel/Brinley', 'Wewanna/Edi', 'Dianne/Hardwell', 'Chadwick Oaks/Clayboard', 'Belleau/Lamar', 'Fetyko/Return', 'Clayboard/Whipporwill Valley', 'Kafka/Milburn', 'Lounga/Herb Hill', 'Packard/Abraham', 'Bernice/N Brookside', 'Grunauer/Michael F', 'Depot/McFeeley', 'Ruby/Katz', 'Parrow/Lamar', 'Whaling/Alize', 'N Martling/Swannekin', 'Deer Hill/Pennybrook', 'N Washington/Soloff', 'Drummond/Stoothoff', 'N Bayview/Dallas', 'Curry/Bergen Hill', 'Fairbury/Slocum', 'Hutter/Brownell', 'Orbach/Wyndover', 'Lange/Yukon', 'Arcadian/Win Haven', 'Corley/Millstone', 'Junker/Turnberry', 'Landry/Hett', 'Grenfell/Clearmeadow', 'Morea/Morea', 'Upper Depew/Motts', 'Klinger/Tornillo', 'Houseman/Beech Hill', 'N Niagra/E Orchard', 'Mc Alpin/N Fulton', 'George/Canarsie', 'Barrie/N Bayview', 'West Lawn/Glennon', 'Meritoria/W Cambridge', 'Snowhill/Aske', 'N Nassau/E Lawn', 'Hadfield/Hagger', 'Sleepy Hollow/Club', 'Booker/W Oxford', 'Fetyko/Motts', 'Sea Shell/Post', 'N Haledon/W Cambridge', 'Black Birch/Jacobs', 'Sperry/Arcadian', 'Mansfield/Ivy Crest', 'E Peddie/Appleby', 'Twin Oaks/Alps', 'Michael F/Pennybrook', 'Hearthwood/Timberline', 'Twin Oak/Aske', 'Handy/Alger', 'Hoyts/W Gilbert', 'Chadwick Oaks/Glennon', 'E Deer Park/Ellin', 'Elzey/Old Field Point', 'Radstock/Loveland', 'Schuman/Easthaven', 'Denham/Lee', 'Flatbush/City Island', 'Samantha/Highhold', 'Allen/Mansfield', 'Ahr/Fortesque', 'Ramapo Mountain/E Deer Park', 'Elmer/N Niagra', 'Chimney Corner/Ellerhausen', 'Yarnell/Alyce', 'Korfitsen/Wylie', 'Meucci/Portside', 'Mansfield/Bixby', 'Milburn/Chedworth', 'Marcus Garvey/Marcus Garvey', 'Dumont/Stivers', 'Jervis/Oxbow', 'Borinsky/Hett', 'Rever/Duda', 'Nicolosi/N Fulton', 'Truro/Newham', 'E Lawn/Bixby', 'Blue Spruce/Wyndover', 'Aske/Motts', 'Brickell/South Airmont', 'Rockaway Valley/Brownell', 'Mc Coy/N Haledon', 'Cedarwood/S Hill', 'E Argyle/Korfitsen', 'N Washington/Sander', 'Foxhunt/Hett', 'Filmore/E Cambridge', 'Schuman/Bergen Hill', 'Overcliff/Easthaven', 'Four Acres/Foxhunt', 'Samantha/Hagger', 'S Wood/Cordwood', 'Fetyko/Chadwick Oaks', 'Tropical/City Island', 'Win Haven/Compass', 'Gladiolus/Korfitsen', 'Tornillo/Milbar', 'Sleepy Hollow/Hanes', 'Bight/Bight', 'Gilburt Hill/Round', 'Win Haven/Ruggles', 'Billington/S Centre Island', 'Banta/Hanes', 'Dreier/Durie', 'Turnberry/Cedarwood', 'East Elbrook/Hellwig', 'Ebersbach/Rockaway Valley', 'Chimney Corner/Embry Farm', 'N Fulton/Roundtop', 'Compass/Gesner', 'Nurge/Leader', 'Newel/Durst', 'Wanamaker/Pool', 'W Penn/Maryann', 'National/National', 'Lynnwood/Maryann', 'Whipporwill Valley/Lebed', 'Midfarm/Sewanee', 'Durst/Stoothoff', 'Round/N Forest', 'Wewanna/Suydam', 'Morris/Stemmer', 'Klinger/S Butehorn', 'Club/Twin Oaks', 'Hardwell/Handy', 'Ringwood/St Andrews', 'Helmetta Jamesburg/W Harris', 'Mount Misery/Micklejohn', 'Pool/Depot', 'Leonardo/Sparkel', 'Edi/Clayboard', 'Arcadian/S Melody', 'Grott/Mc Alpin', 'Schuman/Mc Alpin', 'Bergen Hill/Hampton', 'McFeeley/Balint', 'Dublin/Flagstone', 'Junker/Guntzer', 'Knowles/Black Birch', 'Wyndover/Nipowin', 'Shu Swamp/Pabis', 'Embry Farm/Chedworth', 'Alger/Araca', 'Paddington/Ivy Crest', 'Raybrook/S Sunnyside', 'Ivy Crest/Drummond', 'Filmore/Morris', 'Oakfield/Cheyenne', 'Sperry/Four Acres', 'Cason/Peerless', 'Marginal/Hutter', 'Shirra/Tornillo', 'Post/Millstone', 'Feruzza/Carol', 'Roderick/Melissa', 'Newham/Nipowin', 'Junker/Palisades Center', 'S Beechcroft/Ruggles', 'W Zoranne/Seaton', 'Mount Misery/Feruzza', 'Premier/Sigourney', 'Cliffdale/Bentay', 'Carrie/E Elizabeth', 'Donley/Turnip Hill', 'Dreier/Marne', 'Chippewa/W Hawthorne', 'E Elizabeth/Allan', 'Alize/Weberfield', 'Cason/Embry Farm', 'Grott/Hagger', 'Jada/Meucci', 'Stockholm/Grosbeak', 'Scharer/Red Fern', 'Edi/Grosbeak', 'Cove/Von Huenfeld', 'Mount Misery/Dallas', 'Wycliff/Merrall', 'Nipowin/Esterbrook', 'Premier/Highbrook', 'Elmer/Truro', 'Win Haven/Glennon', 'E Orchard/Lee', 'Lebed/Whipporwill Valley', 'N Haledon/George', 'Hanes/W Cambridge', 'Balint/Merrall', 'Sea Shell/Berriman', 'Whaling/Agatha', 'Depot/E Riviera', 'Weir/Sander', 'Beech Hill/Tropical', 'Korfitsen/Hardwell', 'Pool/St Andrews', 'S Melody/Gary', 'Hagger/Fetyko', 'Leary/Irving', 'Ruggles/Marne', 'Hanes/Reni', 'Shirra/Micklejohn', 'Abington/Hardwell', 'Leeland/Flatbush', 'Scharer/Grenfell', 'Stivers/River Oaks', 'E Orchard/Chippewa', 'Foxhunt/Denham', 'Pilvinis/Windemere', 'Van Tines/Rever', 'Brickell/Ashburton', 'Tailor/Chippewa', 'Hellwig/N Fulton', 'Durie/Bear', 'Bixby/Cos Cob', 'W Linwood/Red Fern', 'Poar/Glennon', 'Grunauer/Hutter', 'Leeland/Round', 'W Oxford/Katz', 'Piercy/Hudson Park', 'Hearth/Old Field Point', 'Heatherwood/Schneider', 'Sperry/S Corona', 'Fiesta/W Zoranne', 'McFeeley/Miro', 'Red Hook/Marne', 'Elzey/Heckelman', 'Filmore/Cheyenne', 'Win Haven/Landry', 'Talcott/Jada', 'Chimney Corner/Esterbrook', 'Bergen Hill/W Linwood', 'Leary/Heckelman', 'Essen/Rockaway Valley', 'W Harris/Ueland', 'Lange/Oakfield', 'S Adelaide/Kashey', 'Orbach/Hiawatha', 'Hagger/Wyndover', 'Pennybrook/Dumont', 'Elzey/Billington', 'Brook Vale/Reith', 'Westgate/Griffith', 'Anchorage/Wewanna', 'W Runyon/Aske', 'Popomora/Balint', 'Dennis/Cason', 'Nilsen/S Oyster Bay', 'N Erie/Heindel', 'Marne/Marcus Garvey', 'Sparkel/Abbott', 'Cherokee/Tiernans', 'Caxton/Berrybush', 'Maolis/Marne', 'Essen/Hadfield', 'Westgate/Lamar', 'S Corona/E Major', 'Twin Oak/Deseret', 'Normalee/Suydam', 'John F Kennedy/Bentay', 'Esterbrook/Bernice', 'Poar/Hanes', 'Paddington/E Inman', 'Lohman/Hagger', 'Bellrose/Mount Misery', 'Bataan/Turnberry', 'Alize/Macombs', 'Carrie/Barrie', 'Samantha/Anchorage', 'Midfarm/Duda', 'Dennis/Old Hill', 'Poar/Packard', 'S Valley/Interlaken', 'Branton/Booker', 'Ellerhausen/Hadfield', 'Hortense/Newel', 'Bellafiore/Ruckman', 'Marne/N Erie', 'Depot/Leary', 'Milliken/Cedarwood', 'Knowles/Hearthwood', 'Lohman/W Gilbert', 'Bellafiore/Katz', 'Wylie/Handy', 'Roberge/Twin Oaks', 'Stony Hill/Pool', 'Stemmer/Newbury', 'S Wood/Abraham', 'Burro/Club', 'Brownell/McKibbin', 'Wostbrock/E Orchard', 'Ashburton/Heckelman', 'Bellafiore/Gary', 'Houseman/S Adelaide', 'Popomora/Chippewa', 'Mount Misery/O Neil', 'Twin Oak/Post', 'East Crescent/Duda', 'N Martling/Currans', 'Bergen Hill/Michigan', 'S Melody/Mc Alpin', 'Oakfield/Brook Vale', 'Hopke/Cutlass', 'Trader/Millstone', 'Hutchison/Lee', 'N Niagra/Lonsdale', 'Blue Spruce/Lohman', 'Lonanbe/Cliffdale', 'Normalee/Dennis', 'Molino/Currans', 'Dennison/W Graham', 'Bond/Deseret', 'Putnum/Saddle Rock', 'S Hill/Wicks', 'Brophy/Roundtop', 'Blue Spruce/E Elizabeth', 'Canarsie/Dixon', 'S Wood/S Beechcroft', 'Donley/Carol', 'Jeanna/Billington', 'E Gate/Timberline', 'Cedarwood/E Broad', 'Jannarone/Leonardo', 'Guntzer/E Argyle', 'Green Ridge/Grunauer', 'Handy/Maryann', 'Micklejohn/Weir', 'Allegany/Herb Hill', 'Roderick/E Runyon', 'Brambach/Hudson Park', 'Booker/Jada', 'New Walnut/Fortesque', 'Yarnell/Premier', 'N Bayview/New Walnut', 'Leader/Whittle', 'Newbury/Loveland', 'Weberfield/East Elbrook', 'Patience/Bentay', 'Minebrook/River Oaks', 'Branton/Chippewa', 'Raybrook/Alkier', 'Oakfield/Christie', 'Hyannis/Cohancy', 'Foxholm/Disposal', 'Upper Depew/Kerrigan', 'Alyce/Bear', 'Stemmer/Hadfield', 'Lanier/E Major', 'E Cambridge/Lohman', 'Hopke/Rolling Field', 'City Island/Alps', 'Booker/Evana', 'Pinto/Hett', 'Bond/S Corona', 'Weir/Cove', 'Embry Farm/Jannarone', 'S Wood/Old Hill', 'Roberge/Bush', 'Gary/Return', 'Club/Leonardo', 'Schneider/Jonel', 'Dennison/Scudder', 'Wilburne/Carneer', 'Lamar/Wedgwood', 'Hyannis/Lynnwood', 'W Graham/Junker', 'Dallas/Fairmead', 'Chimney Corner/Marginal', 'Huntsbridge/Centre View', 'W Penn/Evana', 'City Island/Eagan', 'Old Britton/Jannarone', 'Fiesta/Overcliff', 'Brambach/Wanamaker', 'Jacobs/Ketridge', 'Wedgwood/Ryders', 'Feruzza/Mc Coy', 'Fairmead/Mount Misery', 'Post/Heindel']
    # print(len(street_selector))  # check number of streets

    for vender_iter2 in new_data_set:
        index_iter2 = new_data_set.index(vender_iter2)
        new_data_set[index_iter2][0] = rd.choice(street_selector)



    #####  THE REVISION OF THE LIST WITH NEW, RANDOMIZED DATA IS NOW COMPLETE FOR THIS CITY #####

    ### identify top 5 TIF districts in the city. Could indicate rapidly growing property values and increased taxes within
    ### the district. Used to guide search for companies that may need professional tax help, whether they took subsidies
    ### or not.



    # create list of all TIF districts in the city
    TIF_districts = []
    for district_add in new_data_set:
        if district_add[0] not in TIF_districts:
            TIF_districts.append(district_add[0])

    # create list of all the districts paired with their TIF$ sum
    district_and_sum = []   # this will be the finished list. It will be a list made up of 2-entry lists.
    for district in TIF_districts:   # iterating through TIF district list
        district_index = TIF_districts.index(district)
        varrevised = []   # create an empty list to store district and sum together (for each individual district)
        dis_total = 0.0
        for dist_search in new_data_set:   # iterate through the data set to add up sums $ for each district
            if dist_search[0] == district:
                dis_total += float(dist_search[4])     # ...add it up
        varrevised.append(TIF_districts[district_index])  # append district name to empty list
        varrevised.append(dis_total)   # append TIF$ sum to list as 2nd entry
        district_and_sum.append(varrevised)  # append each 2-entry list to the empty district and sum list


    # change the sums in the district and sum list from strings to integers. Necessary for calculations.
    for intconversion in district_and_sum:
        intconversion[1] = int(intconversion[1])
    if __name__ == "__main__":
        print('The full list of districts with their TIF subsidies is:')
    if __name__ == "__main__":
        for sd in district_and_sum:
            print(sd)
        print()

    # lists to be filled that will make top-5 calculations easier
    pmts_list = []
    top_5_pmts = []
    dist_and_sum_TOP5 = []

    # create list of ONLY the payments from dist and pmt list. It will make it much easier to find top 5 sums.
    for createpmtlist in district_and_sum:
        pmts_list.append(createpmtlist[1])

    # finding top 5 sums in the new payment list using reverse sort method, add to top-5 pmt list
    pmts_list.sort(reverse=True)
    for it in range(5):
        top_5_pmts.append(pmts_list[it])

    if __name__ == "__main__":
        print(f'The top 5 payments are in the amount of : {top_5_pmts}')
        print()

    # match the top 5 amounts to their vendor

    for t5 in top_5_pmts:  # iterate through top 5 pmts
        for check in district_and_sum:  # ...loop to search each of 5 pmts within district and sum list
            index_charlie = district_and_sum.index(check)
            if check[1] == t5: # if a given entry is one of the top 5 payments
                dist_and_sum_TOP5.append(district_and_sum[index_charlie]) #...add the dist/$ to the top 5 list
                break


    ###### FINDING OF THE TOP 5 DISTRICTS AND PAYMENTS IN THIS CITY IS COMPLETE #######





    # Find the vendors who received the top 5 highest TIF money disbursement totals
    generated_vendorlist = []  # previously, the list of vendors for this city was determined but not listed. Newly listed here
    for ven_all in new_data_set:   # add each vendor to generated_vendorlist if not already included
        if ven_all[2] not in generated_vendorlist:
            generated_vendorlist.append(ven_all[2])
    if __name__ == "__main__":
        print('The vendor list generated for this city is:')
        print(generated_vendorlist)

    vendor_and_sum = []  # this will be the finished list. It will be a list made up of 2-entry lists.
    for vend in generated_vendorlist:  # iterating through generated vendor list
        vend_index = generated_vendorlist.index(vend)
        vend_listmaker = []  # create an empty list to store vend and sum together (for each individual vendor)
        vendor_total = 0.0
        for vend_search in new_data_set:  # iterate through the data set to add up sums $ for each vendor
            if vend_search[2] == vend:
                vendor_total += float(vend_search[4])  # ...add it up
        vend_listmaker.append(generated_vendorlist[vend_index])  # append district name to empty list
        vend_listmaker.append(vendor_total)  # append TIF$ sum to list as 2nd entry
        vendor_and_sum.append(vend_listmaker)  # append each 2-entry list to the empty district and sum list


    # change the sums in the district and sum list from strings to integers. Necessary for calculations.
    for intconv in vendor_and_sum:
        intconv[1] = int(intconv[1])
    if __name__ == "__main__":
        print('The full list of VENDORS with their TIF subsidy totals is:')
    if __name__ == "__main__":
        for sd in district_and_sum:
            print(sd)
        print()


    pmts_list2 = []
    top_5_disbursements = []
    vend_and_sum_TOP5 = []

    # create list of ONLY the payments from vend and pmt list. It will make it much easier to find top 5 sums.
    for makepmtlist in vendor_and_sum:
        pmts_list2.append(makepmtlist[1])

    # finding top 5 sums in the new payment list using reverse sort method, add to top-5 pmt list
    pmts_list2.sort(reverse=True)
    for it2 in range(5):
        top_5_disbursements.append(pmts_list2[it2])

    if __name__ == "__main__":
        print(f'The top 5 (vendor total) payments are in the amount of : {top_5_disbursements}')
        print()

    # match the top 5 amounts to their vendor

    for t52 in top_5_disbursements:  # iterate through top 5 pmts
        for check2 in vendor_and_sum:  # ...loop to search each of 5 pmts within vend and sum list
            index_delta = vendor_and_sum.index(check2)
            if check2[1] == t52: # if a given entry is one of the top 5 payments
                vend_and_sum_TOP5.append(vendor_and_sum[index_delta]) #...add the dist/$ to the top 5 list
                break

    if __name__ == "__main__":
        print(f'The top 5 vendors in {city} are in the amount of : {vend_and_sum_TOP5}')
        print()


    ################# CALCULATION OF TOP 5 VENDORS IN THIS CITY IS COMPLETE ############


    # Calculate the sum of all private, real-estate related TIF payments in the city
    tot_TIF_pmts_in_city = 0.0
    for totes in new_data_set:
        tot = float(totes[4])
        tot_TIF_pmts_in_city += tot
    if __name__ == "__main__":
        print(f'The sum of all TIF dollars spent in {city} is ${tot_TIF_pmts_in_city}')


    list_to_return = [tot_TIF_pmts_in_city, vend_and_sum_TOP5, dist_and_sum_TOP5]

    # transfer results to a new, formatted spreadsheet
    # IT WILL CONTAIN THE FINAL LIST FOR THE CITY, THE TOTAL TIF MONEY SPENT, THE TOP 5 DISTRICTS/PMTS, TOP 5 VENDORS
    wb = xl.Workbook()
    ws = wb.active
    ws.merge_cells('A1:H1')
    ws['A1'] = f'TIF District Report: {city}'
    from openpyxl.styles import Alignment

    bottom_border = Border(bottom=Side(style='thin'))
    left_border = Border(left=Side(style='thin'))
    right_border = Border(right=Side(style='thin'))
    top_border = Border(top=Side(style='thin'))


    # creating/formatting title
    TitleCell = ws['A1']
    TitleCell.alignment = Alignment(horizontal='center')
    TitleCell.font = Font(size=36, bold=True)
    # TitleCell.fill = PatternFill(bgColor='0091CB91', fill_type="solid")

    # the formatting of data categories on spreadsheet
    ws['A2'] = 'TIF Dist Name'
    ws['A2'].font = Font(size=18)
    ws['A2'].border = bottom_border
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['B2'] = 'Year'
    ws['B2'].font = Font(size=18)
    ws['B2'].border = bottom_border
    ws['B2'].alignment = Alignment(horizontal='center')
    ws['C2'] = 'Recipient'
    ws['C2'].font = Font(size=18)
    ws['C2'].border = bottom_border
    ws['C2'].alignment = Alignment(horizontal='center')
    ws['D2'] = 'Disbursement Type'
    ws['D2'].font = Font(size=18)
    ws['D2'].border = bottom_border
    ws['D2'].alignment = Alignment(horizontal='center')
    ws['E2'] = 'Payout Amt'
    ws['E2'].font = Font(size=18)
    ws['E2'].border = bottom_border
    ws['E2'].alignment = Alignment(horizontal='center')

    # adding data list to spreadsheet
    for da in new_data_set:
        ws.append(da)

    # setting width of columns for easier readability
    for column in string.ascii_uppercase:
        if (column == 'A'):
            ws.column_dimensions[column].width = 30
        elif (column == 'B'):
            ws.column_dimensions[column].width = 7
        elif (column == 'C'):
            ws.column_dimensions[column].width = 37
        elif (column == 'D'):
            ws.column_dimensions[column].width = 29
        elif (column == 'G'):
            ws.column_dimensions[column].width = 29
        elif (column == 'H'):
            ws.column_dimensions[column].width = 16
        else:
            ws.column_dimensions[column].width = 19

    # using currency formatting for the tif dollars column
    for dol in range(2,400):
        cell = ws[f'E{dol}']
        cell.number_format = '$#,##0_-'   # this is the format code for currency formatting

    ws.merge_cells('G3:H3')
    ws.merge_cells('G4:H4')
    ws.merge_cells('G6:H6')
    ws.merge_cells('G13:H13')

    # create/format total tif payents table
    ws['G3'] = f'Sum of all {city} TIF payments'
    ws['G3'].alignment = Alignment(horizontal='center')
    ws['G3'].font = Font(size=14, bold=True)
    ws['G3'].border = bottom_border
    ws['H3'].border = bottom_border

    ws['G4'] = tot_TIF_pmts_in_city
    ws['G4'].alignment = Alignment(horizontal='center')
    ws['G4'].number_format = '$#,##0_-'

    ws['G6'] = 'Top 5: Subsidies to Vendor'
    ws['G6'].alignment = Alignment(horizontal='center')
    ws['G6'].border = bottom_border
    ws['H6'].border = bottom_border
    ws['G6'].font = Font(size=14, bold=True)
    for num in range(1,6):
        numint = int(num)
        vendor_list_index = numint - 1
        ws[f'G{str(numint + 6)}'] = vend_and_sum_TOP5[vendor_list_index][0]
        ws[f'H{str(numint + 6)}'] = vend_and_sum_TOP5[vendor_list_index][1]
        ws[f'H{str(numint + 6)}'].number_format = '$#,##0_-'


    ws['G13'] = 'Top 5: Subsidies to TIF District'
    ws['G13'].alignment = Alignment(horizontal='center')
    ws['G13'].font = Font(size=14, bold=True)
    ws['G13'].border = bottom_border
    ws['H13'].border = bottom_border
    for num2 in range(1,6):
        numint2 = int(num2)
        dist_list_index2 = numint2 - 1
        ws[f'G{str(numint2 + 13)}'] = dist_and_sum_TOP5[dist_list_index2][0]
        ws[f'H{str(numint2 + 13)}'] = dist_and_sum_TOP5[dist_list_index2][1]
        ws[f'H{str(numint2 + 13)}'].number_format = '$#,##0_-'

    # for column2 in string.ascii_uppercase:
    #     if (column2 == 'G'):
    #         ws.column2_dimensions[column2].width = 29
    #     if (column2 == 'H'):
    #         ws.column2_dimensions[column2].width = 15

    # create unique final name for each city-spreadsheet and save it
    save_destination = f'C:\\Users\\Owner\\Desktop\\TIF sheets\\{city}.xlsx'
    wb.save(save_destination)

    # returns the essential data of this function
    return list_to_return


